import pandas as pd 
import os
import openpyxl

def inv_update(barcodeInput, dir_path, dir_path_db): 
    df = pd.read_excel(dir_path_db, engine='openpyxl', usecols=[3], sheet_name='Sheet1')
    for index, row in df.iterrows():
        if int(barcodeInput) == row.values:
            wb = openpyxl.load_workbook(dir_path_db) 
            sheets = wb.sheetnames
            ws = wb[sheets[0]] 
            filecell = "E" + str(index + 2)
            cell_value = str(ws[filecell].value)
            if cell_value == 'None':
                 ws[filecell].value = 0
                 cell_value = "0"
            print("Current inventory: " + cell_value)
            inv_update_input = input("Enter amount to add or subtract, use negative values to subtract: ")
            inv_update_num = int(cell_value) + int(inv_update_input) 
            print("Updated inventory: " + str(inv_update_num))
            ws[filecell].value = inv_update_num
            wb.save('test2.xlsx')
            wb.close()
            return 
    return False

def main():
    dir_path = os.path.dirname(os.path.realpath(__file__))
    dir_path_db = os.path.join(dir_path, 'test2.xlsx')
    scan = True
    while scan:
        barcodeInput = input("Scan barcode then press enter: ")
        inv_update_found = inv_update(barcodeInput, dir_path, dir_path_db)
        if inv_update_found == False:
            print("Could not find design")
        
main()
