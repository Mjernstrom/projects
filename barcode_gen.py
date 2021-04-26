import kivy
import os
import pandas as pd
from kivy.app import App
from kivy.core.window import Window
from kivy.uix.label import Label
import string
import random
import openpyxl

class MyApp(App):

    def build(self):
        # Create window and add drag and drop property
         Window.bind(on_dropfile=self._on_file_drop)
         label = Label(text='Drag and Drop File Here')
         return label

    def _on_file_drop(self, window, file_path):
        # Get last directory in the path (design file)
        design_file = os.path.basename(os.path.normpath(file_path)).decode('utf-8')
        # Create directory for the excel file
        dir_path_db = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'test.xlsx')
        print(dir_path_db)
        # Open excel sheet and edit barcode row
        df = pd.read_excel(dir_path_db, engine='openpyxl', usecols=[0], sheet_name='Sheet1')
        for index, row in df.iterrows():
            if design_file == row.values:
                letters = string.digits
                barcode = int(( ''.join(random.choice(letters) for i in range(12)) ))
                wb = openpyxl.load_workbook(dir_path_db) 
                sheets = wb.sheetnames
                ws = wb[sheets[0]] 
                filecell = "B" + str(index + 2)
                ws[filecell].value = barcode
                wb.save('test.xlsx')
                wb.close()
                return 
        return
    
MyApp().run()
